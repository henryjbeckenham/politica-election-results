from __future__ import annotations

import csv
import io
import math
import os
import stat
import zipfile
from collections.abc import Iterator
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path, PurePosixPath
from typing import BinaryIO, TextIO

from .detection import schema_signature


SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
MAX_XLSX_EXPANDED_BYTES = 2 * 1024**3
MAX_XLSX_ARCHIVE_MEMBERS = 10_000
MAX_COMPRESSION_RATIO = 500


class InputInspectionError(ValueError):
    pass


def json_value(value):
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        return None if not math.isfinite(value) else value
    if isinstance(value, (datetime, date, Decimal)):
        return str(value)
    return str(value)


def _safe_zip_member(info: zipfile.ZipInfo) -> None:
    member = PurePosixPath(info.filename)
    if member.is_absolute() or ".." in member.parts or "\x00" in info.filename:
        raise InputInspectionError(f"Unsafe ZIP member path: {info.filename}")
    mode = (info.external_attr >> 16) & 0o170000
    if mode == stat.S_IFLNK:
        raise InputInspectionError(f"ZIP symbolic links are not accepted: {info.filename}")


def _validate_xlsx_package(source, virtual_name: str) -> None:
    """Bound the internal ZIP package before openpyxl decompresses it."""

    try:
        package = zipfile.ZipFile(source)
    except zipfile.BadZipFile as exc:
        raise InputInspectionError(f"Invalid XLSX upload: {virtual_name}") from exc
    try:
        members = [info for info in package.infolist() if not info.is_dir()]
        if len(members) > MAX_XLSX_ARCHIVE_MEMBERS:
            raise InputInspectionError(
                f"XLSX contains {len(members):,} package members; "
                f"limit is {MAX_XLSX_ARCHIVE_MEMBERS:,}"
            )
        total_size = sum(info.file_size for info in members)
        if total_size > MAX_XLSX_EXPANDED_BYTES:
            raise InputInspectionError(
                f"XLSX expands to {total_size:,} bytes; "
                f"limit is {MAX_XLSX_EXPANDED_BYTES:,}"
            )
        for info in members:
            _safe_zip_member(info)
            if (
                info.compress_size
                and info.file_size / info.compress_size > MAX_COMPRESSION_RATIO
            ):
                raise InputInspectionError(
                    f"XLSX package member has an unsafe compression ratio: {info.filename}"
                )
    finally:
        package.close()
        if hasattr(source, "seek"):
            source.seek(0)


def _decode_sample(sample: bytes) -> tuple[str, str]:
    try:
        return sample.decode("utf-8-sig"), "utf-8-sig"
    except UnicodeDecodeError:
        return sample.decode("cp1252"), "cp1252"


def _csv_dialect(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        return ","


def _open_csv_text(binary: BinaryIO, encoding: str) -> TextIO:
    return io.TextIOWrapper(binary, encoding=encoding, newline="")


def _csv_header_and_preview(
    opener,
    *,
    preview_rows: int,
) -> tuple[list[str], list[dict], str, str, int]:
    with opener() as raw:
        sample = raw.read(64 * 1024)
    sample_text, encoding = _decode_sample(sample)
    delimiter = _csv_dialect(sample_text)
    with opener() as raw:
        with _open_csv_text(raw, encoding) as text:
            reader = csv.reader(text, delimiter=delimiter)
            try:
                first = next(reader)
            except StopIteration as exc:
                raise InputInspectionError("CSV file is empty") from exc
            header_row_number = 1
            if len(first) == 1 and "Event:" in first[0]:
                try:
                    first = next(reader)
                    header_row_number = 2
                except StopIteration as exc:
                    raise InputInspectionError("CSV contains a preamble but no header row") from exc
            headers = [str(value).strip() for value in first]
            if not headers or not any(headers):
                raise InputInspectionError("CSV header row is empty")
            if len(headers) != len(set(headers)):
                raise InputInspectionError("CSV contains duplicate column headers")
            preview: list[dict] = []
            for values in reader:
                if not values or not any(str(value).strip() for value in values):
                    continue
                padded = values[: len(headers)] + [None] * max(0, len(headers) - len(values))
                preview.append(dict(zip(headers, padded, strict=True)))
                if len(preview) >= preview_rows:
                    break
    return headers, preview, encoding, delimiter, header_row_number


def _xlsx_workbook(source, *, read_only: bool = True):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX support requires the openpyxl package") from exc
    return load_workbook(source, read_only=read_only, data_only=True)


def _worksheet_header(sheet) -> tuple[int, list[str]]:
    for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
        serialised = [json_value(value) for value in values]
        nonempty = [value for value in serialised if value not in (None, "")]
        if not nonempty:
            continue
        if len(nonempty) == 1 and "Event:" in str(nonempty[0]):
            continue
        headers = ["" if value is None else str(value).strip() for value in serialised]
        while headers and not headers[-1]:
            headers.pop()
        if not headers or not any(headers):
            continue
        if any(not header for header in headers):
            raise InputInspectionError(
                f"Worksheet {sheet.title!r} has a blank header between populated columns"
            )
        if len(headers) != len(set(headers)):
            raise InputInspectionError(f"Worksheet {sheet.title!r} has duplicate headers")
        return row_number, headers
    raise InputInspectionError(f"Worksheet {sheet.title!r} has no header row")


def _xlsx_preview(sheet, header_row: int, headers: list[str], limit: int) -> list[dict]:
    preview: list[dict] = []
    for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number <= header_row:
            continue
        serialised = [json_value(value) for value in values[: len(headers)]]
        serialised += [None] * max(0, len(headers) - len(serialised))
        if not any(value not in (None, "") for value in serialised):
            continue
        preview.append(dict(zip(headers, serialised, strict=True)))
        if len(preview) >= limit:
            break
    return preview


def inspect_upload(
    path: Path,
    upload_id: str,
    original_name: str,
    *,
    preview_rows: int,
    max_archive_bytes: int,
    max_archive_members: int,
    max_xlsx_member_bytes: int,
) -> tuple[list[dict], list[dict]]:
    extension = Path(original_name).suffix.lower()
    if extension == ".zip":
        return _inspect_zip(
            path,
            upload_id,
            original_name,
            preview_rows=preview_rows,
            max_archive_bytes=max_archive_bytes,
            max_archive_members=max_archive_members,
            max_xlsx_member_bytes=max_xlsx_member_bytes,
        )
    if extension not in SUPPORTED_EXTENSIONS:
        raise InputInspectionError("Only CSV, XLSX and ZIP uploads are accepted")
    datasets = _inspect_file(
        path,
        upload_id,
        original_name,
        member=None,
        preview_rows=preview_rows,
        max_xlsx_member_bytes=max_xlsx_member_bytes,
    )
    return datasets, []


def _inspect_zip(
    path: Path,
    upload_id: str,
    original_name: str,
    *,
    preview_rows: int,
    max_archive_bytes: int,
    max_archive_members: int,
    max_xlsx_member_bytes: int,
) -> tuple[list[dict], list[dict]]:
    datasets: list[dict] = []
    ignored: list[dict] = []
    try:
        archive = zipfile.ZipFile(path)
    except zipfile.BadZipFile as exc:
        raise InputInspectionError(f"Invalid ZIP upload: {original_name}") from exc
    with archive:
        members = [info for info in archive.infolist() if not info.is_dir()]
        if len(members) > max_archive_members:
            raise InputInspectionError(
                f"ZIP contains {len(members):,} files; limit is {max_archive_members:,}"
            )
        total_size = sum(info.file_size for info in members)
        if total_size > max_archive_bytes:
            raise InputInspectionError(
                f"ZIP expands to {total_size:,} bytes; limit is {max_archive_bytes:,}"
            )
        for info in members:
            _safe_zip_member(info)
            member_extension = Path(info.filename).suffix.lower()
            if member_extension not in SUPPORTED_EXTENSIONS:
                ignored.append(
                    {"member": info.filename, "reason": "unsupported_extension", "size_bytes": info.file_size}
                )
                continue
            if info.compress_size and info.file_size / info.compress_size > MAX_COMPRESSION_RATIO:
                raise InputInspectionError(f"ZIP member has an unsafe compression ratio: {info.filename}")
            datasets.extend(
                _inspect_file(
                    path,
                    upload_id,
                    info.filename,
                    member=info.filename,
                    preview_rows=preview_rows,
                    max_xlsx_member_bytes=max_xlsx_member_bytes,
                )
            )
    if not datasets:
        raise InputInspectionError("ZIP contains no supported CSV or XLSX datasets")
    return datasets, ignored


def _inspect_file(
    container_path: Path,
    upload_id: str,
    virtual_name: str,
    *,
    member: str | None,
    preview_rows: int,
    max_xlsx_member_bytes: int,
) -> list[dict]:
    extension = Path(virtual_name).suffix.lower()
    if extension == ".csv":
        if member:
            def opener():
                archive = zipfile.ZipFile(container_path)
                binary = archive.open(member)

                class ArchiveStream:
                    def __enter__(self):
                        return binary

                    def __exit__(self, *_):
                        binary.close()
                        archive.close()

                return ArchiveStream()
        else:
            def opener():
                return container_path.open("rb")

        headers, preview, encoding, delimiter, header_row = _csv_header_and_preview(
            opener, preview_rows=preview_rows
        )
        return [
            {
                "upload_id": upload_id,
                "virtual_name": virtual_name,
                "member": member,
                "sheet": None,
                "format": "csv",
                "headers": headers,
                "header_row": header_row,
                "preview": preview,
                "encoding": encoding,
                "delimiter": delimiter,
                "schema_signature_sha256": schema_signature(headers),
                "row_count": None,
            }
        ]

    if member:
        with zipfile.ZipFile(container_path) as archive:
            info = archive.getinfo(member)
            if info.file_size > max_xlsx_member_bytes:
                raise InputInspectionError(
                    f"XLSX member {member} is too large to inspect safely ({info.file_size:,} bytes)"
                )
            source = io.BytesIO(archive.read(member))
            _validate_xlsx_package(source, virtual_name)
            workbook = _xlsx_workbook(source)
    else:
        _validate_xlsx_package(container_path, virtual_name)
        workbook = _xlsx_workbook(container_path)
    try:
        datasets = []
        for sheet in workbook.worksheets:
            header_row, headers = _worksheet_header(sheet)
            preview = _xlsx_preview(sheet, header_row, headers, preview_rows)
            if not preview and sheet.max_row <= header_row:
                continue
            datasets.append(
                {
                    "upload_id": upload_id,
                    "virtual_name": virtual_name,
                    "member": member,
                    "sheet": sheet.title,
                    "format": "xlsx",
                    "headers": headers,
                    "header_row": header_row,
                    "preview": preview,
                    "encoding": None,
                    "delimiter": None,
                    "schema_signature_sha256": schema_signature(headers),
                    "row_count": max(0, sheet.max_row - header_row),
                }
            )
        if not datasets:
            raise InputInspectionError(f"XLSX file {virtual_name} has no populated worksheets")
        return datasets
    finally:
        workbook.close()


def iter_dataset_rows(container_path: Path, dataset: dict) -> Iterator[tuple[int, dict]]:
    if dataset["format"] == "csv":
        yield from _iter_csv_rows(container_path, dataset)
        return
    yield from _iter_xlsx_rows(container_path, dataset)


def _iter_csv_rows(container_path: Path, dataset: dict) -> Iterator[tuple[int, dict]]:
    member = dataset.get("member")
    if member:
        archive = zipfile.ZipFile(container_path)
        raw = archive.open(member)
    else:
        archive = None
        raw = container_path.open("rb")
    try:
        with _open_csv_text(raw, dataset["encoding"]) as text:
            reader = csv.reader(text, delimiter=dataset["delimiter"])
            for _ in range(dataset["header_row"]):
                try:
                    next(reader)
                except StopIteration:
                    return
            headers = dataset["headers"]
            for source_row_number, values in enumerate(
                reader, start=dataset["header_row"] + 1
            ):
                if not values or not any(str(value).strip() for value in values):
                    continue
                padded = values[: len(headers)] + [None] * max(0, len(headers) - len(values))
                yield source_row_number, dict(zip(headers, padded, strict=True))
    finally:
        raw.close()
        if archive is not None:
            archive.close()


def _iter_xlsx_rows(container_path: Path, dataset: dict) -> Iterator[tuple[int, dict]]:
    member = dataset.get("member")
    if member:
        with zipfile.ZipFile(container_path) as archive:
            source = io.BytesIO(archive.read(member))
        _validate_xlsx_package(source, dataset["virtual_name"])
        workbook = _xlsx_workbook(source)
    else:
        _validate_xlsx_package(container_path, dataset["virtual_name"])
        workbook = _xlsx_workbook(container_path)
    try:
        sheet = workbook[dataset["sheet"]]
        headers = dataset["headers"]
        for source_row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
            if source_row_number <= dataset["header_row"]:
                continue
            serialised = [json_value(value) for value in values[: len(headers)]]
            serialised += [None] * max(0, len(headers) - len(serialised))
            if not any(value not in (None, "") for value in serialised):
                continue
            yield source_row_number, dict(zip(headers, serialised, strict=True))
    finally:
        workbook.close()
